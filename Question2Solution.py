#Wilcoxan Wilcoxon Signed Rank Test-for all permutations of the non-key columns and Using 100% equal weighting
#calculate the weighted combination of the non-key columns and the run through the Wilkoxon Signed Rank test.

import xlsxwriter
import itertools
import Question1Solution
from openpyxl import load_workbook


def create_combination(filepath):
    file="data/out.xlsx"
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    sheet_max_row = sheet.max_row
    sheet_max_column = sheet.max_column
    counter=2
    dict = {}
    header = []

    #To get combinations of diffrent non-key columns'''
    for i in sheet.iter_rows(min_row=1,max_row=1,min_col=2,max_col=sheet_max_column,values_only=True):
        header.append(i)
    combination=(list(itertools.combinations(header[0],2)))

    #'''to get key-column in dict'''
    col1_values=[]
    col2_values=[]
    for value in sheet.iter_rows(min_row=2, max_row=sheet_max_row, min_col=1, max_col=2, values_only=True):
        col1,col2=value
        print(col1,col2)
        col1_values.append(col1)
        col2_values.append(col2)
    dict.update({"key1":col1_values})
    dict.update({"key2":col2_values})

    ColNames = {}
    Current  = 0
    for COL in sheet.iter_cols(1, sheet.max_column):
        ColNames[COL[0].value] = Current
        Current += 1
    print(ColNames)

    for combo in combination:
        col1, col2 = combo
        value = []
        key = "W_" + str(col1) + "_" + str(col2)
        for row_cells in sheet.iter_rows(min_row=2, max_row=sheet_max_row):
            col1_values=(row_cells[ColNames[col1]].value)
            col2_values=(row_cells[ColNames[col2]].value)
            value.append((col1_values*1+col2_values*1)/2)
        dict.update({key:value})
    return dict

filepath="data/Algo_Test.xlsx"
weighted_dict=create_combination(filepath)

#writting dictionary of weighted combination to xlsx file
file='data/combination_weighted_columns.xlsx'
workbook = xlsxwriter.Workbook(file)
worksheet = workbook.add_worksheet()
col_num = 0
for key, value in weighted_dict.items():
    worksheet.write(0, col_num, key)
    worksheet.write_column(1, col_num, value)
    col_num += 1
workbook.close()

#calling wilixcon signed rank test algoritham for each weighted column
dict_weighted_comibnation=Question1Solution.wilcoxanAlgo(file)
#this prints the column number that best fit to key product
w = min(dict_weighted_comibnation.values())
for k, v in dict_weighted_comibnation.items():
    if v == w:
        print("The best fit to key product is column number ", k)

