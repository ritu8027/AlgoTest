#Wilcoxan Wilcoxon Signed Rank Test
#Use Xslx file as input with key and non-key columns

import rank
from openpyxl import load_workbook

#this function takes ranking list as input and return the summation of
#all ranks having negative diffrences
def sum_negative_differences(sub_list):
    w_minus=0
    for value in sub_list:
        if value[0][1]<0:
            w_minus+=(value[1])
    return w_minus

#this function takes ranking list as input and return the summation of
#all ranks having postive diffrences
def sum_positive_diffrence(sub_list):
    w_plus=0
    for value in sub_list:
        if value[0][1]>0:
            w_plus+=value[1]
    return w_plus

#this function takes xlsx file path with key-non key columns and values ,
#returns the dictionary with key : column number and value of abs(W+ - W-)  that is test statistics for each pair
def wilcoxanAlgo(filepath):
    #loading workbook and activating first sheet
    file="data/out.xlsx"
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    counter = 2                                    #count is taken as 2 as exclueded first 2 key columns
    dict = {}
    sheet_max_row = sheet.max_row                  #maximum row count
    sheet_max_column = sheet.max_column            #maximun column count

    #Iterate through each column and rows
    while(sheet_max_column!=counter):
        diff=[]
        for value in sheet.iter_rows(min_row=2,max_row=sheet_max_row,min_col=1,max_col=3,values_only=True):
            customer,indep_key,depend_key=value                                         #unpacks tuple values
            diff.append([customer,indep_key-depend_key,abs(indep_key-depend_key)])      #taking abs diffrence , normal diffrence ank key in list
        ignored_zero_cases_list = []
        for i in diff:                                                                   #ignoring o diffrences
            if int(i[1])!= 0:
                ignored_zero_cases_list.append(i)
        ranking_data=[]
        for i in ignored_zero_cases_list:
            ranking_data.append(i[2])
        ranks = rank.rankdata(ranking_data)
        data_ranks = list(zip(ignored_zero_cases_list, ranks))
        w_poistive = sum_positive_diffrence(data_ranks)
        w_negative = sum_negative_differences(data_ranks)
        print("Test Statistc W is", min(w_poistive, w_negative))
        w=w_poistive-w_negative
        dict.update({(counter+1):abs(w)})
        sheet.delete_cols(idx=3,amount=1)
        counter+=1
        workbook.save(file)
    return dict


filepath="data/Algo_Test.xlsx"
dict=wilcoxanAlgo(filepath)
#this prints the column number that best fit to key product
w = min(dict.values())
for k, v in dict.items():
    if v == w:
        print("The best fit to key product is column number ", k)
